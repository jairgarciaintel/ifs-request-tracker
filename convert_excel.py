"""
Convert requests.xlsx or requests.csv (exported from SharePoint) to requests.json
Then auto-push to GitHub.

Usage:
    python convert_excel.py
"""
import json
import subprocess
import sys
import csv
from pathlib import Path
from datetime import datetime

REPO_DIR = Path(__file__).parent
DATA_DIR = REPO_DIR / 'data'
EXCEL_FILE = DATA_DIR / 'requests.xlsx'
CSV_FILE = DATA_DIR / 'requests.csv'
JSON_FILE = DATA_DIR / 'requests.json'


def convert():
    # Try CSV first (simpler, no dependency needed), then XLSX
    if CSV_FILE.exists():
        return convert_csv()
    elif EXCEL_FILE.exists():
        return convert_xlsx()
    else:
        print(f"Error: No data file found. Place requests.csv or requests.xlsx in data/")
        sys.exit(1)


def convert_csv():
    rows = []
    with open(str(CSV_FILE), 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for item in reader:
            if not item.get('ID'):
                continue
            rows.append({
                'id': int(item['ID']) if item['ID'].isdigit() else item['ID'],
                'requestType': item.get('Request Type') or 'Unknown',
                'customer': item.get('Company') or item.get('Project/Portal Name') or '',
                'requestor': item.get('Created By') or '',
                'created': str(item.get('Created') or ''),
                'status': item.get('iGO Admin Only - Status') or '',
                'engagementType': item.get('Engagement Type') or '',
                'priority': item.get('Priority') or '',
                'projectName': item.get('Project/Portal Name') or ''
            })
    
    with open(str(JSON_FILE), 'w') as f:
        json.dump(rows, f, indent=2, default=str)
    
    print(f"Converted {len(rows)} requests (CSV) -> {JSON_FILE}")
    return len(rows)


def convert_xlsx():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Use CSV instead, or: pip install openpyxl")
        sys.exit(1)
    
    wb = openpyxl.load_workbook(str(EXCEL_FILE))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        if item.get('ID') is None:
            continue
        rows.append({
            'id': item['ID'],
            'requestType': item.get('Request Type') or 'Unknown',
            'customer': item.get('Company') or item.get('Project/Portal Name') or '',
            'requestor': item.get('Created By') or '',
            'created': str(item.get('Created') or ''),
            'status': item.get('iGO Admin Only - Status') or '',
            'engagementType': item.get('Engagement Type') or '',
            'priority': item.get('Priority') or '',
            'projectName': item.get('Project/Portal Name') or ''
        })
    
    with open(str(JSON_FILE), 'w') as f:
        json.dump(rows, f, indent=2, default=str)
    
    print(f"Converted {len(rows)} requests (XLSX) -> {JSON_FILE}")
    return len(rows)
    return len(rows)


def git_push():
    git = 'git'
    git_win = Path(r'C:\Program Files\Git\bin\git.exe')
    if git_win.exists():
        git = str(git_win)

    # Read token from .env
    env_file = REPO_DIR / '.env'
    token = ''
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if line.startswith('GITHUB_TOKEN='):
                token = line.split('=', 1)[1].strip()

    if token:
        remote = f'https://jairgarciaintel:{token}@github.com/jairgarciaintel/ifs-request-tracker.git'
        subprocess.run([git, 'remote', 'set-url', 'origin', remote], cwd=str(REPO_DIR), capture_output=True)

    subprocess.run([git, 'add', 'data/'], cwd=str(REPO_DIR), capture_output=True)
    
    msg = f'Sync: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    result = subprocess.run([git, 'commit', '-m', msg], cwd=str(REPO_DIR), capture_output=True, text=True)
    
    if 'nothing to commit' in (result.stdout + result.stderr):
        print('No changes to push.')
        return

    push = subprocess.run([git, 'push'], cwd=str(REPO_DIR), capture_output=True, text=True)
    if push.returncode == 0:
        print('Pushed to GitHub. Page updates in ~1 min.')
    else:
        print(f'Push failed: {push.stderr}')


if __name__ == '__main__':
    print("=" * 40)
    print("IFS Request Tracker - Excel to JSON")
    print("=" * 40)
    convert()
    print("\nPushing to GitHub...")
    git_push()
    print(f"\nDone! https://jairgarciaintel.github.io/ifs-request-tracker/")
