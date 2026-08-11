"""
IFS Request Tracker — SharePoint Sync Script
=============================================
Este script:
1. Conecta a la SharePoint List "New DA Request" en intel.sharepoint.com
2. Descarga todos los requests desde julio 1, 2026
3. Guarda en data/requests.json (para la landing page)
4. Guarda en data/requests.xlsx (para referencia en Excel)

Requisitos:
    pip install office365-rest-python-client openpyxl

Uso:
    python sync_sharepoint.py

    O programar cada 10 minutos con cron/Task Scheduler:
    */10 * * * * cd /path/to/Request-Tracker && python sync_sharepoint.py

Autenticación:
    Usa tu cuenta Intel (username + password) o un App Registration.
    La primera vez te pide credenciales y las guarda en .env (gitignored).
"""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

# Output paths
DATA_DIR = Path(__file__).parent / 'data'
DATA_DIR.mkdir(exist_ok=True)
REQUESTS_JSON = DATA_DIR / 'requests.json'
REQUESTS_XLSX = DATA_DIR / 'requests.xlsx'
TRACKING_JSON = DATA_DIR / 'tracking.json'

# SharePoint configuration
SP_SITE = 'https://intel.sharepoint.com/sites/ifs-igo-requests'
SP_LIST = 'New DA Request'
DATE_FILTER = '2026-07-01T00:00:00Z'


def get_credentials():
    """Load credentials from .env or prompt user."""
    env_file = Path(__file__).parent / '.env'
    
    if env_file.exists():
        creds = {}
        for line in env_file.read_text().splitlines():
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                creds[k.strip()] = v.strip()
        return creds.get('SP_USERNAME'), creds.get('SP_PASSWORD')
    
    print("First run — enter your Intel credentials.")
    print("These will be saved to .env (gitignored).\n")
    username = input("Intel email (e.g., jair.garcia@intel.com): ").strip()
    password = input("Password: ").strip()
    
    env_file.write_text(f"# SharePoint credentials (DO NOT COMMIT)\nSP_USERNAME={username}\nSP_PASSWORD={password}\n")
    print(f"✅ Saved to {env_file}\n")
    return username, password


def fetch_from_sharepoint(username, password):
    """Fetch list items from SharePoint REST API."""
    try:
        from office365.runtime.auth.user_credential import UserCredential
        from office365.sharepoint.client_context import ClientContext
    except ImportError:
        print("❌ Missing dependency. Run:")
        print("   pip install office365-rest-python-client")
        sys.exit(1)
    
    print(f"🔗 Connecting to {SP_SITE}...")
    ctx = ClientContext(SP_SITE).with_credentials(UserCredential(username, password))
    
    print(f"📋 Fetching list: {SP_LIST}...")
    sp_list = ctx.web.lists.get_by_title(SP_LIST)
    
    # Build CAML query for items since July 1, 2026
    caml_query = f"""
    <View>
        <Query>
            <Where>
                <And>
                    <Geq>
                        <FieldRef Name='Created'/>
                        <Value Type='DateTime'>{DATE_FILTER}</Value>
                    </Geq>
                    <And>
                        <Neq><FieldRef Name='Status'/><Value Type='Text'>Complete</Value></Neq>
                        <And>
                            <Neq><FieldRef Name='Status'/><Value Type='Text'>Closed</Value></Neq>
                            <Neq><FieldRef Name='Status'/><Value Type='Text'>Hidden</Value></Neq>
                        </And>
                    </And>
                </And>
            </Where>
            <OrderBy><FieldRef Name='ID' Ascending='FALSE'/></OrderBy>
        </Query>
        <RowLimit>1000</RowLimit>
    </View>
    """
    
    from office365.sharepoint.listitems.caml.query import CamlQuery
    qry = CamlQuery()
    qry.ViewXml = caml_query
    
    items = sp_list.get_items(qry)
    ctx.execute_query()
    
    print(f"✅ Retrieved {len(items)} items")
    return items


def normalize_items(items):
    """Convert SharePoint items to clean JSON format."""
    results = []
    for item in items:
        props = item.properties
        results.append({
            'id': props.get('Id') or props.get('ID'),
            'requestType': props.get('Request_x0020_Type') or props.get('RequestType') or props.get('Title') or 'Unknown',
            'customer': props.get('Customer_x0020_Name') or props.get('CustomerName') or props.get('Customer') or '—',
            'requestor': props.get('Author') or '—',
            'created': props.get('Created', ''),
            'status': props.get('Status', '')
        })
    return results


def save_json(requests):
    """Save requests to JSON file."""
    REQUESTS_JSON.write_text(json.dumps(requests, indent=2, default=str))
    print(f"💾 Saved {len(requests)} requests → {REQUESTS_JSON}")


def save_excel(requests):
    """Save requests to Excel file with tracking status merged in."""
    try:
        import openpyxl
    except ImportError:
        print("⚠️  openpyxl not installed — skipping Excel export")
        print("   pip install openpyxl")
        return
    
    # Load existing tracking data
    tracking = {}
    if TRACKING_JSON.exists():
        tracking = json.loads(TRACKING_JSON.read_text())
    
    # Sub-steps definition
    sub_steps = {
        'Portal Creation': ['Configure AGS', 'Configure JSM', 'Configure Security FT', 'Configure Access Mgmt', 'WebView AGS Approver'],
        'New DA': ['Acknowledge', 'In Approval Loop', 'Add Force Signature'],
        'Codename': ['Codename Created'],
        'WebView': ['WebView Done']
    }
    
    sub_step_ids = {
        'Portal Creation': ['configure-ags', 'configure-jsm', 'configure-security-ft', 'configure-access-mgmt', 'webview-ags-approver'],
        'New DA': ['acknowledge', 'in-approval-loop', 'add-force-signature'],
        'Codename': ['codename-created'],
        'WebView': ['webview-done']
    }
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Request Tracker'
    
    # Header
    headers = ['ID', 'Customer', 'Request Type', 'Requestor', 'Date', 'Sub-Step', 'Status', 'Last Updated']
    ws.append(headers)
    
    # Style header
    from openpyxl.styles import Font, PatternFill
    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='0071C5', end_color='0071C5', fill_type='solid')
    
    # Data rows
    for req in requests:
        key = f"{req['id']}-{req['requestType']}"
        steps = sub_steps.get(req['requestType'], [])
        step_ids = sub_step_ids.get(req['requestType'], [])
        saved = tracking.get(key, {})
        
        if not steps:
            ws.append([req['id'], req['customer'], req['requestType'], req['requestor'], req['created'], '—', 'N/A', ''])
        else:
            for label, sid in zip(steps, step_ids):
                status = saved.get(sid, 'Not Started')
                ws.append([req['id'], req['customer'], req['requestType'], req['requestor'], req['created'], label, status, saved.get('_updated', '')])
    
    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)
    
    wb.save(str(REQUESTS_XLSX))
    print(f"📊 Saved Excel → {REQUESTS_XLSX}")


def git_push():
    """Commit and push updated data to GitHub so the page refreshes."""
    import subprocess
    
    repo_dir = Path(__file__).parent
    git = 'git'
    
    # Try to find git on Windows if not in PATH
    git_win = Path(r'C:\Program Files\Git\bin\git.exe')
    if git_win.exists():
        git = str(git_win)
    
    try:
        # Configure remote with token (reads from .env)
        env_file = repo_dir / '.env'
        token = ''
        if env_file.exists():
            for line in env_file.read_text().splitlines():
                if line.startswith('GITHUB_TOKEN='):
                    token = line.split('=', 1)[1].strip()
        
        if token:
            remote_url = f'https://jairgarciaintel:{token}@github.com/jairgarciaintel/ifs-request-tracker.git'
            subprocess.run([git, 'remote', 'set-url', 'origin', remote_url], cwd=repo_dir, capture_output=True)
        
        # Add, commit, push
        subprocess.run([git, 'add', 'data/'], cwd=repo_dir, capture_output=True)
        
        msg = f'Sync: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        result = subprocess.run([git, 'commit', '-m', msg], cwd=repo_dir, capture_output=True, text=True)
        
        if 'nothing to commit' in result.stdout:
            print('   No changes to push (data unchanged).')
            return
        
        push_result = subprocess.run([git, 'push'], cwd=repo_dir, capture_output=True, text=True)
        
        if push_result.returncode == 0:
            print('🚀 Pushed to GitHub — page will update in ~1 min.')
        else:
            print(f'⚠️  Push failed: {push_result.stderr}')
    except Exception as e:
        print(f'⚠️  Git push failed: {e}')
        print('   Data saved locally. Push manually with: git add data/ && git commit -m "sync" && git push')


def main():
    print("=" * 50)
    print("IFS Request Tracker — SharePoint Sync")
    print(f"Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 50)
    
    username, password = get_credentials()
    
    try:
        items = fetch_from_sharepoint(username, password)
        requests = normalize_items(items)
    except Exception as e:
        print(f"\n❌ SharePoint connection failed: {e}")
        print("   Using existing local data if available.")
        if REQUESTS_JSON.exists():
            requests = json.loads(REQUESTS_JSON.read_text())
            print(f"   Loaded {len(requests)} from cache.")
        else:
            print("   No cache available. Exiting.")
            sys.exit(1)
    
    save_json(requests)
    save_excel(requests)
    
    # Auto-push to GitHub so the page updates
    print("\n📤 Pushing to GitHub...")
    git_push()
    
    print(f"\n✅ Done! Page: https://jairgarciaintel.github.io/ifs-request-tracker/")


if __name__ == '__main__':
    main()
