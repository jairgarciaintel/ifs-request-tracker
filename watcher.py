"""
FS Request Tracker — File Watcher Bot
======================================
Watches for new/updated .iqy or .xlsx files in the data/ folder.
When detected:
1. If .iqy → opens it to refresh Excel data (you still need to Save As .xlsx manually)
2. If .xlsx → converts to JSON and pushes to GitHub automatically

Install:
    pip install watchdog openpyxl

Usage:
    python watcher.py

Leave it running in the background. When you download the .iqy from SharePoint
and save it as .xlsx in the data/ folder, it auto-converts and pushes.
"""

import time
import json
import subprocess
import sys
from pathlib import Path
from datetime import datetime

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    print("Missing dependency. Run:")
    print("  pip install watchdog")
    sys.exit(1)

REPO_DIR = Path(__file__).parent
DATA_DIR = REPO_DIR / 'data'
XLSX_FILE = DATA_DIR / 'requests.xlsx'
JSON_FILE = DATA_DIR / 'requests.json'


def convert_and_push():
    """Convert xlsx to json and push to GitHub."""
    print(f"\n{'='*40}")
    print(f"[{datetime.now().strftime('%H:%M:%S')}] Change detected!")
    print(f"{'='*40}")
    
    if not XLSX_FILE.exists():
        print("  No requests.xlsx found. Waiting...")
        return
    
    # Convert
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(XLSX_FILE))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = dict(zip(headers, row))
            if item.get('ID') is None:
                continue
            rows.append({
                'id': item['ID'],
                'requestType': item.get('Request Type') or 'Unknown',
                'customer': item.get('Company') or item.get('Project/Portal Name') or '',
                'requestor': item.get('Created By') or '',
                'created': str(item.get('Created') or ''),
                'status': item.get('iGO Admin Only - Status') or '',
                'engagementType': item.get('Engagement Type') or '',
                'priority': item.get('Priority') or '',
                'projectName': item.get('Project/Portal Name') or ''
            })
        
        with open(str(JSON_FILE), 'w') as f:
            json.dump(rows, f, indent=2, default=str)
        
        print(f"  Converted {len(rows)} requests -> requests.json")
    except Exception as e:
        print(f"  Conversion error: {e}")
        return
    
    # Git push
    git = 'git'
    git_win = Path(r'C:\Program Files\Git\bin\git.exe')
    if git_win.exists():
        git = str(git_win)
    
    try:
        subprocess.run([git, 'add', 'data/requests.json'], cwd=str(REPO_DIR), capture_output=True)
        msg = f'Auto-sync: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
        result = subprocess.run([git, 'commit', '-m', msg], cwd=str(REPO_DIR), capture_output=True, text=True)
        
        if 'nothing to commit' in (result.stdout + result.stderr):
            print('  No changes in data (same as before).')
            return
        
        push = subprocess.run([git, 'push', 'origin', 'main'], cwd=str(REPO_DIR), capture_output=True, text=True)
        if push.returncode == 0:
            print('  Pushed to GitHub! Page updates in ~1 min.')
        else:
            print(f'  Push failed: {push.stderr[:200]}')
    except Exception as e:
        print(f'  Git error: {e}')


class XlsxHandler(FileSystemEventHandler):
    def __init__(self):
        self.last_trigger = 0
    
    def on_modified(self, event):
        self._handle(event)
    
    def on_created(self, event):
        self._handle(event)
    
    def _handle(self, event):
        if event.is_directory:
            return
        path = Path(event.src_path)
        # Only react to .xlsx files
        if path.suffix.lower() != '.xlsx':
            return
        # Debounce: ignore events within 5 seconds of each other
        now = time.time()
        if now - self.last_trigger < 5:
            return
        self.last_trigger = now
        # Wait a moment for file to finish writing
        time.sleep(2)
        convert_and_push()


def main():
    DATA_DIR.mkdir(exist_ok=True)
    
    print("="*50)
    print("FS Request Tracker — File Watcher Bot")
    print("="*50)
    print(f"\n  Watching: {DATA_DIR}")
    print(f"  Trigger:  Save any .xlsx file here")
    print(f"\n  When you download from SharePoint:")
    print(f"  1. Open the .iqy in Excel")
    print(f"  2. File > Save As > {XLSX_FILE}")
    print(f"  3. This bot auto-converts and pushes!")
    print(f"\n  Press Ctrl+C to stop.\n")
    
    handler = XlsxHandler()
    observer = Observer()
    observer.schedule(handler, str(DATA_DIR), recursive=False)
    observer.start()
    
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
        print("\n  Watcher stopped.")
    observer.join()


if __name__ == '__main__':
    main()
